#!/usr/bin/env python3
"""
excel_exporter.py
Enhanced Excel export functionality for warehouse space measurement data.
Creates structured Excel reports with analysis, charts, and insights.
"""

import pandas as pd
import numpy as np
import requests
import json
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

class WarehouseExcelExporter:
    def __init__(self, api_base_url="http://localhost:5001"):
        self.api_base_url = api_base_url
        self.output_dir = "exports"
        os.makedirs(self.output_dir, exist_ok=True)
        
    def fetch_warehouse_data(self):
        """Fetch warehouse data from API"""
        try:
            response = requests.get(f"{self.api_base_url}/api/data", timeout=10)
            if response.status_code == 200:
                return response.json()
            else:
                print(f"⚠️  API response error: {response.status_code}")
                return []
        except Exception as e:
            print(f"❌ Error fetching data: {e}")
            return []
    
    def create_comprehensive_excel_report(self, data=None):
        """Create comprehensive Excel report with multiple sheets"""
        if data is None:
            data = self.fetch_warehouse_data()
        
        if not data:
            print("⚠️  No data available for export")
            return None
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Raw Data
        self._create_raw_data_sheet(wb, df)
        
        # Sheet 2: Summary Analysis
        self._create_summary_sheet(wb, df)
        
        # Sheet 3: Capacity Analysis
        self._create_capacity_analysis_sheet(wb, df)
        
        # Sheet 4: Condition Report
        self._create_condition_report_sheet(wb, df)
        
        # Sheet 5: Recommendations
        self._create_recommendations_sheet(wb, df)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"warehouse_space_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb.save(filepath)
        print(f"✅ Excel report saved: {filepath}")
        return filepath
    
    def _create_raw_data_sheet(self, wb, df):
        """Create raw data sheet with all warehouse information"""
        ws = wb.create_sheet(title="Raw Data")
        
        # Headers
        headers = [
            "ID", "Aisle", "Rack ID", "Rack Type", "Total Capacity", 
            "Occupied Space", "Free Space", "Condition", "Notes", "Timestamp",
            "Occupancy %", "Usable Area (sqm)", "Accessibility Score"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row.get('id', ''))
            ws.cell(row=row_idx, column=2, value=row.get('aisle', ''))
            ws.cell(row=row_idx, column=3, value=row.get('rackId', ''))
            ws.cell(row=row_idx, column=4, value=row.get('rackType', ''))
            ws.cell(row=row_idx, column=5, value=row.get('totalCapacity', 0))
            ws.cell(row=row_idx, column=6, value=row.get('occupiedSpace', 0))
            ws.cell(row=row_idx, column=7, value=row.get('freeSpace', 0))
            ws.cell(row=row_idx, column=8, value=row.get('condition', ''))
            ws.cell(row=row_idx, column=9, value=row.get('notes', ''))
            ws.cell(row=row_idx, column=10, value=row.get('timestamp', ''))
            
            # Calculate occupancy percentage
            total_cap = row.get('totalCapacity', 1)
            occupied = row.get('occupiedSpace', 0)
            occupancy_pct = (occupied / total_cap * 100) if total_cap > 0 else 0
            ws.cell(row=row_idx, column=11, value=f"{occupancy_pct:.1f}%")
            
            # Additional metrics (if available)
            ws.cell(row=row_idx, column=12, value=row.get('usable_area_sqm', 'N/A'))
            ws.cell(row=row_idx, column=13, value=row.get('accessibility_score', 'N/A'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb, df):
        """Create summary analysis sheet"""
        ws = wb.create_sheet(title="Summary Analysis")
        
        # Title
        ws['A1'] = "Warehouse Space Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Summary statistics
        total_racks = len(df)
        total_capacity = df['totalCapacity'].sum()
        total_occupied = df['occupiedSpace'].sum()
        total_free = df['freeSpace'].sum()
        avg_occupancy = (total_occupied / total_capacity * 100) if total_capacity > 0 else 0
        
        # Create summary table
        summary_data = [
            ["Metric", "Value"],
            ["Total Racks", total_racks],
            ["Total Capacity (units)", total_capacity],
            ["Currently Occupied (units)", total_occupied],
            ["Available Space (units)", total_free],
            ["Average Occupancy", f"{avg_occupancy:.1f}%"],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Aisle breakdown
        ws['A11'] = "Breakdown by Aisle"
        ws['A11'].font = Font(size=14, bold=True)
        
        aisle_summary = df.groupby('aisle').agg({
            'totalCapacity': 'sum',
            'occupiedSpace': 'sum',
            'freeSpace': 'sum'
        }).reset_index()
        
        aisle_headers = ["Aisle", "Total Capacity", "Occupied", "Free Space", "Occupancy %"]
        for col_idx, header in enumerate(aisle_headers, 1):
            cell = ws.cell(row=12, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for row_idx, (_, row) in enumerate(aisle_summary.iterrows(), 13):
            ws.cell(row=row_idx, column=1, value=row['aisle'])
            ws.cell(row=row_idx, column=2, value=row['totalCapacity'])
            ws.cell(row=row_idx, column=3, value=row['occupiedSpace'])
            ws.cell(row=row_idx, column=4, value=row['freeSpace'])
            occupancy = (row['occupiedSpace'] / row['totalCapacity'] * 100) if row['totalCapacity'] > 0 else 0
            ws.cell(row=row_idx, column=5, value=f"{occupancy:.1f}%")
        
        # Add chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Capacity by Aisle"
        chart.y_axis.title = 'Units'
        chart.x_axis.title = 'Aisle'
        
        data = Reference(ws, min_col=2, min_row=12, max_col=4, max_row=12 + len(aisle_summary))
        cats = Reference(ws, min_col=1, min_row=13, max_row=12 + len(aisle_summary))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
    
    def _create_capacity_analysis_sheet(self, wb, df):
        """Create capacity analysis sheet"""
        ws = wb.create_sheet(title="Capacity Analysis")
        
        # Title
        ws['A1'] = "Warehouse Capacity Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Utilization categories
        df['utilization_category'] = df.apply(lambda row: self._categorize_utilization(
            row['occupiedSpace'], row['totalCapacity']), axis=1)
        
        utilization_counts = df['utilization_category'].value_counts()
        
        # Create utilization table
        ws['A3'] = "Space Utilization Distribution"
        ws['A3'].font = Font(size=14, bold=True)
        
        util_headers = ["Utilization Level", "Count", "Percentage"]
        for col_idx, header in enumerate(util_headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        
        total_racks = len(df)
        row_idx = 5
        for category, count in utilization_counts.items():
            percentage = (count / total_racks * 100) if total_racks > 0 else 0
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.1f}%")
            row_idx += 1
        
        # Efficiency metrics
        ws['A11'] = "Efficiency Metrics"
        ws['A11'].font = Font(size=14, bold=True)
        
        # Calculate efficiency metrics
        underutilized = len(df[df['utilization_category'] == 'Underutilized'])
        optimal = len(df[df['utilization_category'] == 'Optimal'])
        overutilized = len(df[df['utilization_category'] == 'Overutilized'])
        
        efficiency_data = [
            ["Metric", "Value", "Status"],
            ["Space Efficiency", f"{(optimal/total_racks*100):.1f}%", "Good" if optimal/total_racks > 0.6 else "Needs Improvement"],
            ["Underutilization Risk", f"{(underutilized/total_racks*100):.1f}%", "Low" if underutilized/total_racks < 0.3 else "High"],
            ["Capacity Strain", f"{(overutilized/total_racks*100):.1f}%", "Low" if overutilized/total_racks < 0.1 else "High"]
        ]
        
        for row_idx, row_data in enumerate(efficiency_data, 12):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
    
    def _create_condition_report_sheet(self, wb, df):
        """Create condition report sheet"""
        ws = wb.create_sheet(title="Condition Report")
        
        # Title
        ws['A1'] = "Warehouse Condition Assessment"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Condition summary
        condition_counts = df['condition'].value_counts()
        
        ws['A3'] = "Condition Distribution"
        ws['A3'].font = Font(size=14, bold=True)
        
        condition_headers = ["Condition", "Count", "Percentage"]
        for col_idx, header in enumerate(condition_headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        
        total_racks = len(df)
        row_idx = 5
        for condition, count in condition_counts.items():
            percentage = (count / total_racks * 100) if total_racks > 0 else 0
            ws.cell(row=row_idx, column=1, value=condition)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.1f}%")
            
            # Color code based on condition
            if condition == 'good':
                fill_color = "90EE90"  # Light green
            elif condition == 'partially-blocked':
                fill_color = "FFD700"  # Gold
            elif condition == 'damaged':
                fill_color = "FFA07A"  # Light salmon
            else:
                fill_color = "FFFFFF"  # White
            
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row_idx += 1
        
        # Issues requiring attention
        problematic_racks = df[df['condition'] != 'good']
        
        if len(problematic_racks) > 0:
            ws[f'A{row_idx + 2}'] = "Racks Requiring Attention"
            ws[f'A{row_idx + 2}'].font = Font(size=14, bold=True)
            
            issue_headers = ["Rack ID", "Aisle", "Condition", "Notes"]
            for col_idx, header in enumerate(issue_headers, 1):
                cell = ws.cell(row=row_idx + 3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
            
            start_row = row_idx + 4
            for idx, (_, rack) in enumerate(problematic_racks.iterrows()):
                ws.cell(row=start_row + idx, column=1, value=rack['rackId'])
                ws.cell(row=start_row + idx, column=2, value=rack['aisle'])
                ws.cell(row=start_row + idx, column=3, value=rack['condition'])
                ws.cell(row=start_row + idx, column=4, value=rack['notes'])
    
    def _create_recommendations_sheet(self, wb, df):
        """Create recommendations sheet"""
        ws = wb.create_sheet(title="Recommendations")
        
        # Title
        ws['A1'] = "Operational Recommendations"
        ws['A1'].font = Font(size=16, bold=True)
        
        recommendations = self._generate_recommendations(df)
        
        row_idx = 3
        for category, recs in recommendations.items():
            ws[f'A{row_idx}'] = category
            ws[f'A{row_idx}'].font = Font(size=14, bold=True)
            row_idx += 1
            
            for rec in recs:
                ws[f'B{row_idx}'] = f"• {rec}"
                row_idx += 1
            
            row_idx += 1  # Add space between categories
    
    def _categorize_utilization(self, occupied, total):
        """Categorize utilization level"""
        if total == 0:
            return "Unknown"
        
        utilization = occupied / total
        if utilization < 0.3:
            return "Underutilized"
        elif utilization < 0.8:
            return "Optimal"
        else:
            return "Overutilized"
    
    def _generate_recommendations(self, df):
        """Generate operational recommendations based on data analysis"""
        recommendations = {
            "Space Optimization": [],
            "Maintenance Priority": [],
            "Operational Efficiency": [],
            "Capacity Planning": []
        }
        
        # Space optimization
        underutilized = df[df.apply(lambda row: (row['occupiedSpace'] / row['totalCapacity']) < 0.3, axis=1)]
        if len(underutilized) > 0:
            recommendations["Space Optimization"].append(
                f"Consider consolidating items from {len(underutilized)} underutilized racks"
            )
        
        overutilized = df[df.apply(lambda row: (row['occupiedSpace'] / row['totalCapacity']) > 0.9, axis=1)]
        if len(overutilized) > 0:
            recommendations["Space Optimization"].append(
                f"Redistribute items from {len(overutilized)} near-capacity racks"
            )
        
        # Maintenance priority
        damaged_racks = df[df['condition'] == 'damaged']
        if len(damaged_racks) > 0:
            recommendations["Maintenance Priority"].append(
                f"Immediate attention required for {len(damaged_racks)} damaged racks"
            )
        
        blocked_racks = df[df['condition'] == 'partially-blocked']
        if len(blocked_racks) > 0:
            recommendations["Maintenance Priority"].append(
                f"Clear obstructions from {len(blocked_racks)} partially blocked racks"
            )
        
        # Operational efficiency
        total_capacity = df['totalCapacity'].sum()
        total_free = df['freeSpace'].sum()
        efficiency = (total_capacity - total_free) / total_capacity if total_capacity > 0 else 0
        
        if efficiency < 0.6:
            recommendations["Operational Efficiency"].append(
                "Overall warehouse utilization is low - consider space consolidation"
            )
        elif efficiency > 0.9:
            recommendations["Operational Efficiency"].append(
                "Warehouse is near capacity - consider expansion or reorganization"
            )
        
        # Capacity planning
        aisle_utilization = df.groupby('aisle').apply(
            lambda x: x['occupiedSpace'].sum() / x['totalCapacity'].sum()
        ).to_dict()
        
        high_util_aisles = [aisle for aisle, util in aisle_utilization.items() if util > 0.85]
        if high_util_aisles:
            recommendations["Capacity Planning"].append(
                f"Monitor capacity in high-utilization aisles: {', '.join(high_util_aisles)}"
            )
        
        return recommendations
    
    def create_daily_summary_export(self):
        """Create a simple daily summary export"""
        data = self.fetch_warehouse_data()
        if not data:
            return None
        
        df = pd.DataFrame(data)
        
        # Create simple Excel with daily summary
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"daily_warehouse_summary_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Main data
            df.to_excel(writer, sheet_name='Warehouse Data', index=False)
            
            # Summary
            summary_df = pd.DataFrame({
                'Metric': ['Total Racks', 'Total Capacity', 'Occupied Space', 'Free Space', 'Average Occupancy %'],
                'Value': [
                    len(df),
                    df['totalCapacity'].sum(),
                    df['occupiedSpace'].sum(),
                    df['freeSpace'].sum(),
                    f"{(df['occupiedSpace'].sum() / df['totalCapacity'].sum() * 100):.1f}%" if df['totalCapacity'].sum() > 0 else "0%"
                ]
            })
            summary_df.to_excel(writer, sheet_name='Daily Summary', index=False)
        
        print(f"✅ Daily summary exported: {filepath}")
        return filepath

def main():
    """Main function to run Excel export"""
    print("📊 Warehouse Excel Export System")
    print("=" * 40)
    
    exporter = WarehouseExcelExporter()
    
    # Create comprehensive report
    print("🔄 Creating comprehensive Excel report...")
    report_path = exporter.create_comprehensive_excel_report()
    
    if report_path:
        print(f"✅ Report created successfully: {report_path}")
    else:
        print("❌ Failed to create report")
    
    # Create daily summary
    print("🔄 Creating daily summary...")
    summary_path = exporter.create_daily_summary_export()
    
    if summary_path:
        print(f"✅ Daily summary created: {summary_path}")

if __name__ == "__main__":
    main()
